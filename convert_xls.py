#!/usr/bin/env python3
"""
Convertește XLS în XLSX
Folosire: python convert_xls.py input.xls [output.xlsx]
"""
import sys
import os

try:
    import xlrd
    import openpyxl
except ImportError:
    print("Instalează librăriile: pip install xlrd openpyxl")
    sys.exit(1)

def convert_xls_to_xlsx(input_file, output_file=None):
    if not os.path.exists(input_file):
        print(f"Fișierul nu există: {input_file}")
        return False

    if output_file is None:
        output_file = os.path.splitext(input_file)[0] + '.xlsx'

    print(f"Citesc: {input_file}")

    # Citește XLS
    xls_book = xlrd.open_workbook(input_file)

    # Creează XLSX
    xlsx_book = openpyxl.Workbook()
    xlsx_book.remove(xlsx_book.active)  # Remove default sheet

    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name[:31])  # Max 31 chars

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)

        print(f"  Sheet '{xls_sheet.name}': {xls_sheet.nrows} rânduri, {xls_sheet.ncols} coloane")

    # Salvează XLSX
    xlsx_book.save(output_file)
    print(f"Salvat: {output_file}")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Folosire: python convert_xls.py input.xls [output.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if convert_xls_to_xlsx(input_file, output_file):
        print("\n✓ Conversie reușită! Acum poți importa fișierul XLSX.")
    else:
        print("\n✗ Conversie eșuată!")
        sys.exit(1)
